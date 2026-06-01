import io
import openpyxl

def extract_texts(file_bytes: bytes) -> list[dict]:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    segments = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if (
                    isinstance(cell.value, str)
                    and cell.value.strip()
                    and not cell.value.startswith("=")
                ):
                    segments.append({
                        "text": cell.value,
                        "key": (sheet_name, cell.row, cell.column),
                    })
    return segments

def reinsert_texts(file_bytes: bytes, segments: list[dict], translated: list[str]) -> bytes:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    key_to_translation = {
        tuple(s["key"]): t for s, t in zip(segments, translated)
    }
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                key = (sheet_name, cell.row, cell.column)
                if key in key_to_translation:
                    cell.value = key_to_translation[key]
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
