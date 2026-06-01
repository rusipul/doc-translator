import io
import openpyxl
from processors.xlsx import extract_texts, reinsert_texts

def _make_xlsx(cells: list) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    for row_idx, row in enumerate(cells, 1):
        for col_idx, val in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def test_extract_text_cells_only():
    data = _make_xlsx([["안녕", 123, "=A1+1", None, "테스트"]])
    segments = extract_texts(data)
    texts = [s["text"] for s in segments]
    assert "안녕" in texts
    assert "테스트" in texts
    assert 123 not in texts
    assert "=A1+1" not in texts

def test_reinsert_preserves_non_text():
    data = _make_xlsx([[42, "안녕"]])
    segments = extract_texts(data)
    result = reinsert_texts(data, segments, ["Hello"])
    wb = openpyxl.load_workbook(io.BytesIO(result))
    ws = wb.active
    assert ws.cell(1, 1).value == 42
    assert ws.cell(1, 2).value == "Hello"
